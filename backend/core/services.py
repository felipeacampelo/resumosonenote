"""
Serviços de negócio do sistema.

Contém lógica de importação, exportação e processamento de dados.
"""

import openpyxl
from django.db import transaction
from .models import Disciplina, Assunto, Subassunto


class MatrizImportService:
    """
    Serviço para importação da matriz de assuntos a partir de arquivo Excel.
    
    Formato esperado do Excel (Matriz de Assuntos):
    - Múltiplas abas, uma por disciplina
    - Linha 1: vazia
    - Linha 2: cabeçalho
    - Linha 3: vazia
    - Linha 4+: dados
    
    Colunas:
    - Coluna A (1): Assunto
    - Coluna B (2): Subassunto 1 (opcional)
    - Coluna C (3): Subassunto 2 (opcional)
    - Coluna D (4): Aula dos Resumos OneNote (link)
    - Coluna E (5): Caderno TEC Concursos - Cebraspe (link)
    - Coluna G (7): Caderno de Questões FGV (link)
    - Coluna H (8): Dica
    """
    
    def __init__(self):
        self.erros = []
        self.avisos = []
        self.estatisticas = {
            'disciplinas_criadas': 0,
            'assuntos_criados': 0,
            'subassuntos_criados': 0,
            'linhas_processadas': 0,
            'linhas_ignoradas': 0
        }
    
    def importar_arquivo(self, arquivo_path):
        """
        Importa matriz de assuntos de um arquivo Excel.
        
        Args:
            arquivo_path (str): Caminho do arquivo Excel
            
        Returns:
            dict: Estatísticas da importação
            
        Raises:
            ValueError: Se o arquivo for inválido
        """
        try:
            workbook = openpyxl.load_workbook(arquivo_path, data_only=True)
        except Exception as e:
            raise ValueError(f"Erro ao abrir arquivo Excel: {str(e)}")
        
        # Processar cada aba como uma disciplina
        with transaction.atomic():
            for ordem, sheet_name in enumerate(workbook.sheetnames, start=1):
                self._processar_aba(workbook[sheet_name], ordem)
        
        return {
            'sucesso': len(self.erros) == 0,
            'estatisticas': self.estatisticas,
            'erros': self.erros,
            'avisos': self.avisos
        }
    
    def _processar_aba(self, sheet, ordem):
        """
        Processa uma aba do Excel como uma disciplina.
        
        Args:
            sheet: Planilha do openpyxl
            ordem (int): Ordem da disciplina
        """
        nome_disciplina = sheet.title
        
        # Criar ou obter disciplina
        disciplina, criada = Disciplina.objects.get_or_create(
            nome=nome_disciplina,
            defaults={'ordem': ordem, 'ativa': True}
        )
        
        if criada:
            self.estatisticas['disciplinas_criadas'] += 1
        
        # Processar linhas da planilha (começando da linha 4, pois 1=vazia, 2=cabeçalho, 3=vazia)
        ordem_assunto = 0
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            self.estatisticas['linhas_processadas'] += 1
            
            # Ignorar linhas vazias
            if not any(row):
                self.estatisticas['linhas_ignoradas'] += 1
                continue
            
            # Ignorar linhas que são cabeçalhos (contém "Assunto" na primeira coluna)
            primeira_coluna = self._limpar_texto(row[0]) if row[0] else None
            if primeira_coluna and primeira_coluna.lower() == 'assunto':
                self.estatisticas['linhas_ignoradas'] += 1
                continue
            
            nome_assunto = primeira_coluna
            nome_subassunto1 = self._limpar_texto(row[1]) if len(row) > 1 and row[1] else None
            nome_subassunto2 = self._limpar_texto(row[2]) if len(row) > 2 and row[2] else None
            
            # Extrair metadados das colunas
            link_resumos = self._limpar_texto(row[3]) if len(row) > 3 and row[3] else ''
            link_questoes_cebraspe = self._limpar_texto(row[4]) if len(row) > 4 and row[4] else ''
            link_questoes_fgv = self._limpar_texto(row[6]) if len(row) > 6 and row[6] else ''
            dica = self._limpar_texto(row[7]) if len(row) > 7 and row[7] else ''
            
            # Se tem assunto, criar/atualizar
            if nome_assunto:
                ordem_assunto += 1
                
                assunto, criado = Assunto.objects.get_or_create(
                    disciplina=disciplina,
                    nome=nome_assunto,
                    defaults={
                        'ordem': ordem_assunto, 
                        'ativo': True,
                        'link_resumos': link_resumos or '',
                        'link_questoes_cebraspe': link_questoes_cebraspe or '',
                        'link_questoes_fgv': link_questoes_fgv or '',
                        'dica': dica or '',
                    }
                )
                
                if criado:
                    self.estatisticas['assuntos_criados'] += 1
                else:
                    # Atualizar metadados se o assunto já existe
                    atualizado = False
                    if link_resumos and not assunto.link_resumos:
                        assunto.link_resumos = link_resumos
                        atualizado = True
                    if link_questoes_cebraspe and not assunto.link_questoes_cebraspe:
                        assunto.link_questoes_cebraspe = link_questoes_cebraspe
                        atualizado = True
                    if link_questoes_fgv and not assunto.link_questoes_fgv:
                        assunto.link_questoes_fgv = link_questoes_fgv
                        atualizado = True
                    if dica and not assunto.dica:
                        assunto.dica = dica
                        atualizado = True
                    if atualizado:
                        assunto.save()
                
                # Criar subassunto 1 se existir
                if nome_subassunto1:
                    sub1, criado = Subassunto.objects.get_or_create(
                        assunto=assunto,
                        nome=nome_subassunto1,
                        defaults={'ordem': 1, 'ativo': True}
                    )
                    if criado:
                        self.estatisticas['subassuntos_criados'] += 1
                
                # Criar subassunto 2 se existir
                if nome_subassunto2:
                    sub2, criado = Subassunto.objects.get_or_create(
                        assunto=assunto,
                        nome=nome_subassunto2,
                        defaults={'ordem': 2, 'ativo': True}
                    )
                    if criado:
                        self.estatisticas['subassuntos_criados'] += 1
            else:
                # Linha sem assunto mas com subassunto - avisar
                if nome_subassunto1 or nome_subassunto2:
                    self.avisos.append(
                        f"Linha {row_num} em '{nome_disciplina}': "
                        f"Subassunto sem assunto pai"
                    )
    
    def _limpar_texto(self, texto):
        """
        Limpa e normaliza texto.
        
        Args:
            texto: Texto a ser limpo
            
        Returns:
            str: Texto limpo ou None
        """
        if not texto:
            return None
        
        texto = str(texto).strip()
        return texto if texto else None
    
    def limpar_matriz_existente(self):
        """
        Remove toda a matriz existente do banco de dados.
        
        ATENÇÃO: Esta operação é irreversível!
        """
        with transaction.atomic():
            Subassunto.objects.all().delete()
            Assunto.objects.all().delete()
            Disciplina.objects.all().delete()


class ExportacaoTutoryService:
    """
    Serviço para exportação de mapas no formato Tutory.
    
    Formato de exportação (19 colunas):
    1. Disciplina
    2. Assunto
    3. Páginas ou Minutos de Vídeo
    4. Minutos Expresso
    5. Minutos Regular
    6. Minutos Calma
    7. Dica
    8. Dica de Revisões
    9. Dica de Questões
    10. Referência
    11. Ordenação
    12. Peso de Resumos
    13. Peso de Revisões
    14. Peso de Questões
    15. Número de Questões
    16. Link de Estudo
    17. Link de Resumo
    18. Link de Questões
    19. Suplementar
    """
    
    def __init__(self):
        self.workbook = None
        self.sheet = None
    
    def exportar_concurso(self, concurso):
        """
        Exporta um concurso para o formato Tutory.
        
        Args:
            concurso: Objeto Concurso a ser exportado
            
        Returns:
            BytesIO: Arquivo Excel em memória
        """
        from io import BytesIO
        from .models import MapaAssunto
        
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = concurso.sigla[:31]  # Limite de 31 chars para nome de aba
        
        # Cabeçalhos - Formato Tutory
        headers = [
            'Disciplina',
            'Assunto',
            'Páginas ou Minutos de Vídeo',
            'Minutos Expresso',
            'Minutos Regular',
            'Minutos Calma',
            'Dica',
            'Dica de Revisões',
            'Dica de Questões',
            'Referência',
            'Ordenação',
            'Peso de Resumos',
            'Peso de Revisões',
            'Peso de Questões',
            'Número de Questões',
            'Link de Estudo',
            'Link de Resumo',
            'Link de Questões',
            'Suplementar'
        ]
        
        for col, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=col, value=header)
        
        # Dados
        mapas = MapaAssunto.objects.filter(
            concurso=concurso
        ).select_related(
            'assunto', 'assunto__disciplina', 'subassunto'
        ).prefetch_related('metadados').order_by('ordem')
        
        row = 2
        for mapa in mapas:
            assunto = mapa.assunto
            metadados = getattr(mapa, 'metadados', None)
            
            # Disciplina
            disciplina_nome = assunto.disciplina.nome if assunto else ''
            
            # Nome do assunto (com subassunto se houver)
            if mapa.extra_cursinho:
                assunto_nome = mapa.nome_extra
            elif mapa.subassunto:
                assunto_nome = f"{assunto.nome} - {mapa.subassunto.nome}"
            else:
                assunto_nome = assunto.nome if assunto else ''
            
            # Preencher linha com dados do metadados (formato Tutory)
            self.sheet.cell(row=row, column=1, value=disciplina_nome)
            self.sheet.cell(row=row, column=2, value=assunto_nome)
            self.sheet.cell(row=row, column=3, value=metadados.paginas_minutos if metadados else 0)
            self.sheet.cell(row=row, column=4, value=float(metadados.minutos_expresso) if metadados else 0)
            self.sheet.cell(row=row, column=5, value=float(metadados.minutos_regular) if metadados else 0)
            self.sheet.cell(row=row, column=6, value=float(metadados.minutos_calma) if metadados else 0)
            self.sheet.cell(row=row, column=7, value=metadados.dica or '' if metadados else '')
            self.sheet.cell(row=row, column=8, value=metadados.dica_revisoes or '' if metadados else '')
            self.sheet.cell(row=row, column=9, value=metadados.dica_questoes or '' if metadados else '')
            self.sheet.cell(row=row, column=10, value=metadados.referencia or '' if metadados else '')
            self.sheet.cell(row=row, column=11, value=mapa.ordem)
            self.sheet.cell(row=row, column=12, value=metadados.peso_resumos if metadados else 1)
            self.sheet.cell(row=row, column=13, value=metadados.peso_revisoes if metadados else 1)
            self.sheet.cell(row=row, column=14, value=metadados.peso_questoes if metadados else 1)
            self.sheet.cell(row=row, column=15, value=metadados.numero_questoes if metadados else 0)
            self.sheet.cell(row=row, column=16, value=metadados.link_estudo or '' if metadados else '')
            self.sheet.cell(row=row, column=17, value=metadados.link_resumo or '' if metadados else '')
            self.sheet.cell(row=row, column=18, value=metadados.link_questoes or '' if metadados else '')
            self.sheet.cell(row=row, column=19, value=1 if (metadados and metadados.suplementar) else 0)
            
            row += 1
        
        # Salvar em memória
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
